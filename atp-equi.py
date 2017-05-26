#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# ATP-EQUI - Programa para gerar equivalentes de rede para o software ATP
#     Copyright (C) 2015  Rafael Eduardo S. Ristow (resristow)
#     Desenvolvido inicialmente em Eletrosul Centrais Elétricas S/A

#     This program is free software: you can redistribute it and/or modify
#     it under the terms of the GNU General Public License as published by
#     the Free Software Foundation, either version 3 of the License, or
#     (at your option) any later version.

#     This program is distributed in the hope that it will be useful,
#     but WITHOUT ANY WARRANTY; without even the implied warranty of
#     MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#     GNU General Public License for more details.

#     You should have received a copy of the GNU General Public License
#     along with this program.  If not, see <http://www.gnu.org/licenses/>.

import textos
from pathlib import Path
from math import *
from subprocess import run
from sys import exit
import datetime
from openpyxl import load_workbook
import json
import codecs
import gettext
# Traduzir o argparse para o português
gettext.gettext = textos.traduzArgParse
import argparse
from re import search
import pickle as pk
import time
from copy import deepcopy

# x.y.z
# x = major change
# y = Minor change
# z = bugfix
VERSION = "1.1.0"

class args_Handler():
    """Classe para administrar os argumentos de entrada do programa em linha de
    comando."""
    def __init__(self):
        parser = argparse.ArgumentParser(description='Cria um equivalente para o ATP',
                        prog='ATP-EQUI')

        group = parser.add_mutually_exclusive_group()
        parser.add_argument("json",help="Arquivo com as configurações - modelo opcoes.json")
        parser.add_argument("-e",action="store_true",help="imprime equivalente.lib")
        parser.add_argument("-s",action="store_true",help="imprime source.lib")
        group.add_argument("-R",action="store_true",help='''compara niveis de curto entre Anafas e ATP, rodando
        automaticamente os curto-circuitos no ATP''')
        group.add_argument("-Rj",action="store_true",help="mesmo que R, mas não roda o ATP")
        group.add_argument("-b",action="store_true",help="lista barras com equivalentes")
        parser.add_argument("-i",action="store_true",help="imprime .lib da rede interna")
        parser.add_argument("-m",action="store_true", help="configura os arquivos para manobras.")

        self.args = parser.parse_args()

class Nodes:
    """Coleção de barras do equivalente.
    As barras/nós são representados pela Classe 'node'. Aqui elas são concentradas
    num dicionário, 'nodes'.
    """
    def __init__(self):
        self.nodes = {}
        self.repATP = set()

    def addNode(self, node):
        self.nodes[node.numAna] = node

    def get_nomeAna(self, numAna):
        return self.nodes[numAna].nomeAna

    def get_vBase(self, numAna):
        return self.nodes[numAna].vBase

    def get_nomeAtp(self, numAna):
        return self.nodes[numAna].nomeAtp

    def get_nomeGerAtp(self, numAna):
        return self.nodes[numAna].nomeGerAtp

    def get_repATP(self):
        return self.repATP

    def alter(self, numAna = 0, dado = '', attr='nomeAna'):
        if dado != '':
            if attr == 'nomeAna': self.nodes[numAna].nomeAna = dado
            if attr == 'Vbase': self.nodes[numAna].Vbase = dado
            if attr == 'nomeAtp': self.nodes[numAna].nomeAtp = dado
            if attr == 'nomeGerAtp': self.nodes[numAna].nomeGerAtp = dado
            if attr == 'AutoNamed': self.nodes[numAna].AutoNamed = dado

    def get_all(self):
        return self.nodes.values()

    def check_repGerATP(self):
        repGerATP = set()

        for barra in self.get_all():
            gerName = self.get_nomeGerAtp(barra.numAna)
            if gerName != '':
                if gerName in repGerATP:
                    self.alter(numAna = barra.numAna, attr = 'nomeGerAtp', dado = 'F' + gerName[:-1])
                    self.check_repGerATP()
                repGerATP.add(gerName)

    def check_repATP(self):
        tempRep = set()

        for barra in self.get_all():
            atpName = self.get_nomeAtp(barra.numAna)
            if atpName != '':
                if atpName in tempRep:
                    self.repATP.add(atpName)
                tempRep.add(atpName)


class node:
    """
    Classe interna para definir os nos/barras do equivalente.
    """
    def __init__(self, linha=''):
        self.numAna = 0
        self.vBase = 0.0
        self.nomeAna = ''
        self.nomeAtp = ''
        self.nomeGerAtp = ''
        self.volt = 1
        self.ang = 0
        self.aux = 0
        self.AutoNamed = 0
        self.addLinha(linha=linha)

    def addLinha(self, linha):
        "Identifica e separa os valores da linha do DBAR do .ANA"
        if linha != '':
            self.numAna = int(linha[0:5])
            try: self.vBase = float(linha[31:35])
            except(ValueError): pass
            self.nomeAna = linha[9:21].strip()
            try:
                self.aux = int(linha[7])
            except(ValueError):
                self.aux = 0


class Branches:
    def __init__(self):
        self.branches = {}
        self.negs = {'equiv':[], 'interna':[]}

    def addBranch(self, branch, dbar):
        # Acrescenta no dicionário 'self.branches' o novo circuito (ou ramo)
        if not branch.interno:
            for node in branch.nodes:
                if dbar.nodes[node].aux:
                    return 1

        self.branches[branch.nodes] = branch

        # Averigua se há algum valor negativo e acrescenta na lista 'self.negs'
        for param in branch.params.values():
            if param < 0:
                if not branch.interno:
                    self.negs['equiv'].append(branch.nodes)
                else:
                    self.negs['interna'].append(branch.nodes)

        # Calcula os valores em Ohm
        for n in branch.paramsOhm:
            if branch.params[n] == 9999.99:
                branch.paramsOhm[n] = 999999
            else:
                branch.paramsOhm[n] = specialFloat(branch.params[n] *
                        dbar.get_vBase(branch.nodes[0])**2/10000)


    def get_equiNodes(self, inner = 0):
        barrasEquiv = set()

        for n in self.branches.values():
            for node in range(2):
                if inner:
                    if n.interno:
                        barrasEquiv.add(n.nodes[node])
                else:
                    if not n.interno:
                        barrasEquiv.add(n.nodes[node])

        return sorted(barrasEquiv)

    def get_tipo(self, node):
        return self.branches[node].tipo

    def get_all(self, inner = 0):
        saida = []
        for n in self.branches.values():
            if inner:
                if n.interno:
                    saida.append(n)
            else:
                if not n.interno:
                    saida.append(n)
        return saida


class branch:
    """
    Classe que representa um circuito. Os atributos da classe são:
    nodes: tuple com os nós DE e PARA
    tipo: G(erador), (T)rafo
    params: parâmetros em %
    paramsOhm: parâmetros em Ohms
    Inicialmente o circuito é criado com parâmetros zeros, até ser solicitado a
    inclusão dos valores obtidos do arquivo .ANA (addLinha)
    """


    def __init__(self, linha, inner = 0):
        self.nodes = (0,0)
        self.tipo = ''
        self.params = {'r1':0, 'x1':0, 'r0':0, 'x0':0}
        self.paramsOhm = {'r1':0, 'x1':0, 'r0':0, 'x0':0}
        self.addLinha(linha=linha)
        self.interno = inner

    def addLinha(self, linha):
        "Identifica e separa os valores da linha de equivalente do DLIN do .ANA"
        if linha != '':
            self.nodes = int(linha[0:5]), int(linha[7:12])

            try:
                if '.' in linha[17:23]: self.params['r1'] = float(linha[17:23])
                else: self.params['r1'] = float(linha[17:23])/100
            except(ValueError):self.params['r1'] = 0

            try:
                if '.' in linha[23:29]: self.params['x1'] = float(linha[23:29])
                else: self.params['x1'] = float(linha[23:29])/100
            except(ValueError): self.params['x1'] = 0

            try:
                if '.' in linha[29:35]: self.params['r0'] = float(linha[29:35])
                else: self.params['r0'] = float(linha[29:35])/100
            except(ValueError): self.params['r0'] = 0

            try:
                if '.' in linha[35:41]: self.params['x0'] = float(linha[35:41])
                else: self.params['x0'] = float(linha[35:41])/100
            except(ValueError): self.params['x0'] = 0

            if (self.nodes[1] == 0) and (self.params['r1'] != 9999.99):
                self.tipo = 'G'
            else: self.tipo = linha[16]


class specialFloat(float):
    """Classe herdada de float para poder mudar o comportamento do metodo
     __str__, que controla como eh feita a conversao do float para string"""
    def __str__ (self):

        valor = self.__float__()

        if valor > 0:
            if log10(valor) < -4:
                temp = round(valor * 10 ** -floor(log10(valor)))
                return str(temp) + '.e' + str(floor(log10(valor)))

            elif log10(valor) >= 6:
                temp = round(valor * 10 ** -floor(log10(valor)),1)
                return str(temp) + 'e' + str(floor(log10(valor)))


            else: return str(valor)

        elif valor == 0:
            return '0.0'

        else:
            if log10(abs(valor)) < -4: return '0.0'
            else: return str(valor)


# FUNÇÕES ---------------------------------------------------------------------#

def get_DBAR(arquivo, colecao):
    """Obtém todo o conteúdo do código DBAR do arquivo .ANA"""

    flag = 0
    dbar = colecao

    #Acrescenta a barra/nó Terra ao conjunto de barras
    dbar.addNode(node())
    dbar.alter(dado = 'Terra')

    for linha in arquivo:
        if (flag == 3) & ('99999' in linha):
            flag = 0
            break
        if (flag == 1) and not ('(' in linha[0]):
            flag = flag | 2
        if flag == 3:
            dbar.addNode(node(linha))
        if 'dbar' in linha[0:4].lower():
            flag = flag | 1

    return dbar

def get_EQUIV(arquivo, colecao, dbar):
    """Obtém todo o conteúdo dos circuitos de equivalentes (marcados com 998) do
    arquivo .ANA."""

    flag = 0
    equiv = colecao

    for linha in arquivo:
        if flag:
            if '99999' in linha[0:6]:
                break
            if '(' not in linha[0]:
                if '998' in linha[65:75]:
                    equiv.addBranch(branch(linha), dbar)
                else:
                    equiv.addBranch(branch(linha, inner = 1), dbar)
        if ('dlin' in linha[0:4].lower()) or ('dcir' in linha[0:4].lower()):
            flag = flag | 1
    return equiv

def getAtpNames(arqPaths, dbar, dlin, base):
    """Faz a leitura do arquivo de texto com os nomes das barras de fronteira
    para o ATP.
    Esses nomes são acrescentados nas instâncias dos nós.
    Em seguida, é feito uma verificação para ver se não há nomes de nó repeti-
    dos, que é uma desgraça para o ATP.
    Por fim, cria-se uma sequência de nomes de nó com geradores para o ATP.
    Também é feita uma verificacao se ha algum nome de noh para o ATP faltante.
    """

    wb = load_workbook(str(arqPaths['Nomes']))

    ws = wb.worksheets[0]

    # Verifica qual escolha de base de nomes o usuário optou
    if base.lower() == 'epe': offset = 1
    if base.lower() == 'ons': offset = 2

    tabela_Nomes = {}

    for linha in ws.iter_rows(row_offset=1):
        try:
            num_barra = int(linha[offset].value)
        except(TypeError):
            num_barra = None
        nome_barra = linha[3].value
        nome_ATP = linha[0].value

        # Verificação se tem alguma célula vazia, o que invalida toda a linha
        if num_barra == None or nome_barra == None or nome_ATP == None: continue

        tabela_Nomes[num_barra] = [nome_barra, nome_ATP]

    # Leitura do arquivo com valores de tensão e ângulo das fontes
    valsFontes = {}
    try:
        arqFontes = arqPaths['initSource'].open('r')
        for linha in arqFontes:
            if linha[0] != '#':
                temp = linha.split()
                valsFontes[temp[0]] = [float(temp[1]), float(temp[2])]
    except(FileNotFoundError):
        pass

    # Verificacao de falta de nome de no para o ATP (somente para circuitos equivalentes)
    autoEquiv = []
    autoIntern = []
    temp = 1

    for barra in dlin.get_equiNodes(inner = 0)[1:]:
        try:
            dbar.alter(numAna = barra, dado = tabela_Nomes[barra][1], attr = 'nomeAtp')
        except(KeyError):
            autoName = str(temp) + 'Y' * (5 - len(str(temp)))
            dbar.alter(numAna = barra, dado = autoName, attr = 'nomeAtp')
            temp += 1
            autoEquiv.append((barra, dbar.get_nomeAna(barra), autoName))
            dbar.alter(numAna = barra, dado = 1, attr = 'AutoNamed')
        try:
            if dlin.get_tipo((barra,0)) == 'G':
                dbar.alter(numAna = barra, dado = 'F' + dbar.get_nomeAtp(barra)[:-1], attr = 'nomeGerAtp')
        except: pass

    # Verificacao de falta de nome de no para o ATP (rede interna)


    for barra in dlin.get_equiNodes(inner = 1)[1:]:
        try:
            dbar.alter(numAna = barra, dado = tabela_Nomes[barra][1], attr = 'nomeAtp')
        except(KeyError):
            autoName = str(temp) + 'Y' * (5 - len(str(temp)))
            dbar.alter(numAna = barra, dado = autoName, attr = 'nomeAtp')
            temp += 1
            autoIntern.append((barra, dbar.get_nomeAna(barra), autoName))
            dbar.alter(numAna = barra, dado = 1, attr = 'AutoNamed')
        try:
            if dlin.get_tipo((barra,0)) == 'G':
                dbar.alter(numAna = barra, dado = 'F' + tabela_Nomes[barra][1][:-1], attr = 'nomeGerAtp')
        except: pass


    dbar.check_repGerATP()

    dbar.check_repATP()

    ATPnames = {}
    for node in dbar.get_all():
        ATPnames[node.nomeAtp] = node

    for fonte in valsFontes:
        try:
            ATPnames[fonte].volt = valsFontes[fonte][0]
            ATPnames[fonte].ang = valsFontes[fonte][1]
        except(KeyError):
            pass

    return autoEquiv, autoIntern

def makeLib(arqPaths, dlin, dbar, inner = 0):
    """Funçaõ para compor o arquivo-cartão /BRANCH com extensão .lib, do ATP,
    que irá conter a rede equivalentada pelo Anafas."""

    if not inner:
        arquivo = arqPaths['cwd'].resolve() / Path(arqPaths['Ana'].name[0:-4] + '-equivalentes.lib')
    else:
        arquivo = arqPaths['cwd'].resolve() / Path(arqPaths['Ana'].name[0:-4] + '-interna.lib')
    arquivo = arquivo.open('w')

    numTrf = 1
    limTrf = 999
    contTrf = 0

    def strNumTrf(numTrf, contTrf):
        if contTrf:
            return chr(contTrf + 64) + str(numTrf)
        else:
            return str(numTrf)


    arquivo.write('/BRANCH\n')

    if inner:
        branches = dlin.get_all(inner = 1)
    else:
        branches = dlin.get_all(inner = 0)

    # Loop 'for' sobre cada circuito do conjunto DLIN
    for branch in branches:
        # Escreve o delimitador superior, com vários '=' e, embaixo, o nome das
        #barras DE e PARA
        arquivo.write('C ' + 77*'=' + '\n' + 'C =====' +
            dbar.get_nomeAna(branch.nodes[0]) + ' - ' +
            dbar.get_nomeAna(branch.nodes[1]) + '\n')

        # Compões a lista com o nome do nó DE para cada uma das 3 fases
        nodeFrom = [str(dbar.get_nomeAtp(branch.nodes[0]))+'A',
                    str(dbar.get_nomeAtp(branch.nodes[0]))+'B',
                    str(dbar.get_nomeAtp(branch.nodes[0]))+'C']

        # A seguir, é feita a composição da lista com o nome do nó PARA, a de-
        #pender do tipo de elemento que se deseja representar no ATP


        # Se for um 'Gerador':
        if branch.tipo == 'G':
            gerName = dbar.get_nomeGerAtp(branch.nodes[0])
            nodeTo = [str(gerName)+'A',
                    str(gerName)+'B',
                    str(gerName)+'C']

        # Se for um 'Shunt':
        elif branch.nodes[1] == 0:
            nodeTo = [' '*6 for n in range(3)]

        # Se for um 'Trafo':
        elif branch.tipo == 'T':
             nodeTo = [strNumTrf(numTrf, contTrf)+'TIA',
                        strNumTrf(numTrf, contTrf)+'TIB',
                        strNumTrf(numTrf, contTrf)+'TIC']

        # Se for uma 'LT':
        else:
            nodeTo = [str(dbar.get_nomeAtp(branch.nodes[1]))+'A',
                    str(dbar.get_nomeAtp(branch.nodes[1]))+'B',
                    str(dbar.get_nomeAtp(branch.nodes[1]))+'C']

        # A seguir, é feita a escrita dos dados do circuito no arquivo-cartão.

        arquivo.write('51{}{:6}'.format(nodeFrom[0],nodeTo[0]) + 12*' ' +
            '{0!s:<6.6}'.format(branch.paramsOhm['r0']) + 6*' ' +
            '{0!s:<12.12}'.format(branch.paramsOhm['x0']) + '\n')
        arquivo.write('52{}{:6}'.format(nodeFrom[1],nodeTo[1]) + 12*' ' +
            '{0!s:<6.6}'.format(branch.paramsOhm['r1']) + 6*' ' +
            '{0!s:<12.12}'.format(branch.paramsOhm['x1']) + '\n')
        arquivo.write('53' + nodeFrom[2] + nodeTo[2] + '\n')

        # Se for um transformador, é colocado um transformador ideal para fazer
        #a relação de tensão entre as duas barras. A impedância da ligação já é
        #escrita na etapa acima.

        if branch.tipo == 'T':


            arquivo.write('  TRANSFORMER' + 25*' ' + 'ATI'+
                '{0:<3s}'.format(strNumTrf(numTrf, contTrf))+'1.E6\n')
            arquivo.write(' '*12 + '9999' + '\n')
            arquivo.write(' 1'+'{0:<6s}'.format(strNumTrf(numTrf, contTrf)+'TIA') + ' '*18 +
                         '.00001.00001'+str(dbar.get_vBase(branch.nodes[0])) + '\n')
            arquivo.write(' 2'+ dbar.get_nomeAtp(branch.nodes[1]) +'A' + ' '*18 +
                         '.00001.00001'+str(dbar.get_vBase(branch.nodes[1]))+ '\n')

            arquivo.write('  TRANSFORMER ATI'+'{0:<3s}'.format(strNumTrf(numTrf, contTrf)) +
                        ' '*18 + '{0:<6s}'.format('BTI'+strNumTrf(numTrf, contTrf))+'\n')
            arquivo.write(' 1'+'{0:<6s}'.format(strNumTrf(numTrf, contTrf) + 'TIB') + '\n')
            arquivo.write(' 2'+dbar.get_nomeAtp(branch.nodes[1])+'B' + '\n')

            arquivo.write('  TRANSFORMER ATI'+'{0:<3s}'.format(strNumTrf(numTrf, contTrf)) +
                        ' '*18 + '{0:<6s}'.format('CTI'+strNumTrf(numTrf, contTrf))+'\n')
            arquivo.write(' 1'+'{0:<6s}'.format(strNumTrf(numTrf, contTrf) + 'TIC') + '\n')
            arquivo.write(' 2'+dbar.get_nomeAtp(branch.nodes[1])+'C' + '\n')

            if numTrf > (limTrf - 1):
                numTrf = 0
                limTrf = 98
                contTrf += 1

            numTrf += 1

def makeSource(arqPaths, dbar):
    """Escreve um arquivo-cartão /SOURCE no formato .lib com as fontes do siste-
    equivalente."""
    arquivo = arqPaths['cwd'].resolve() / Path(arqPaths['Ana'].name[0:-4] + '-source.lib')
    arquivo = arquivo.open('w')
    arquivo.write('/SOURCE\nC < n 1><>< Ampl.  >< Freq.  ><Phase/T0><   A1   ><   T1   >< TSTART >< TSTOP  >\n')
    for barra in dbar.get_all():
        if barra.nomeGerAtp != '':
            arquivo.write('14{:5}A  {!s:10.10}{!s:>10.10}{!s:>10.10}'.format(barra.nomeGerAtp,
                barra.vBase*sqrt(2/3)*1000 * barra.volt, 60, barra.ang) + 20*' ' +
                '{:10d}{:10d}'.format(-1,100) + '\n')
            arquivo.write('14{:5}B  {!s:10.10}{!s:>10.10}{!s:>10.10}'.format(barra.nomeGerAtp,
                barra.vBase*sqrt(2/3)*1000 * barra.volt, 60, -120 + barra.ang) + 20*' ' +
                '{:10d}{:10d}'.format(-1,100) + '\n')
            arquivo.write('14{:5}C  {!s:10.10}{!s:>10.10}{!s:>10.10}'.format(barra.nomeGerAtp,
                barra.vBase*sqrt(2/3)*1000 * barra.volt, 60, -240 + barra.ang) + 20*' ' +
                '{:10d}{:10d}'.format(-1,100) + '\n')

def compCurto(arqPaths, dbar, jump=0):
    """Descrição aqui
    """

    valsCurto = {}

    rncc = arqPaths['Rncc'].open('r')
    flag = 0

    for linha in rncc:
        if flag:
            if 'X-----X------------X' in linha:
                break

            numAna = int(linha[2:7])
            nome = linha[8:20]
            vBase = linha[21:26]
            valsCurto[numAna] = {'3F':{'ANAFAS':[float(linha[28:37]), float(linha[46:53])]}}
            valsCurto[numAna]['1F'] = {'ANAFAS':[float(linha[54:63]), float(linha[71:79])]}


        if 'X-----X------------X' in linha:
            flag = 1
            rncc.readline()

    rncc.close()


    atpWork = arqPaths['ATPcwd']

    if not jump:
        batchPath = atpWork / Path('rodaTodos.bat')
        tempBatch = batchPath.open('w')

        tempBatch.write('set GNUDIR=' + str(arqPaths['GNUDIR']) + '\\\n\n')
        tempBatch.write('cd /d ' + str(batchPath.parent) + '\n\n')

        batchConst = [str(arqPaths['tpbigDir'] / arqPaths['tpbig']) + ' disk ', ' s -r\n']


        flag = 'ok'

        for barra in dbar.get_all():

            # if (barra.numAna == 0) or barra.AutoNamed:
            if (barra.numAna == 0) or (barra.numAna not in valsCurto.keys()):
                continue

            for tipo in ['3F', '1F']:

                baseAtp = arqPaths['Atp'].open('r')

                tempAtp = atpWork / Path(arqPaths['Atp'].stem + tipo + '_' + str(barra.numAna) + '.atp')

                with tempAtp.open('w') as arqAtp:

                    for linha in baseAtp:

                        if flag == 'top':
                            arqAtp.write(linha[:8] + 6 * ' ' + '-1' + linha[16:])
                            flag = 'ok'
                            continue

                        if flag == 'branch':
                            arqAtp.write('C ===RESISTÊNCIA DE CURTO-CIRCUITO===\n')
                            arqAtp.write('  ' +  'CURTOA' + 18 * ' ' + '1e-5\n')
                            arqAtp.write('  ' +  'CURTOB' + 18 * ' ' + '1e-5\n')
                            arqAtp.write('  ' +  'CURTOC' + 18 * ' ' + '1e-5\n')
                            flag = 'ok'

                        if 'switch' in flag:
                            arqAtp.write('C ===LIGAÇÃO COM O CURTO===\n')
                            arqAtp.write('C CORRESPONDE A BARRA ' + str(barra.numAna) + '\n')
                            arqAtp.write('  ' + barra.nomeAtp + 'A' + 'CURTOA' + 7 * ' ' + '-1.' + 6 * ' ' + '1.E3' + 45 * ' ' + '1\n')
                            if tipo == '3F':
                                arqAtp.write('  ' + barra.nomeAtp + 'B' + 'CURTOB' + 7 * ' ' + '-1.' + 6 * ' ' + '1.E3' + 45 * ' ' + '1\n')
                                arqAtp.write('  ' + barra.nomeAtp + 'C' + 'CURTOC' + 7 * ' ' + '-1.' + 6 * ' ' + '1.E3' + 45 * ' ' + '1\n')

                            else:
                                arqAtp.write('  ' + barra.nomeAtp + 'B' + 'CURTOB' + 7 * ' ' + '10.' + 6 * ' ' + '1.E3' + 45 * ' ' + '1\n')
                                arqAtp.write('  ' + barra.nomeAtp + 'C' + 'CURTOC' + 7 * ' ' + '10.' + 6 * ' ' + '1.E3' + 45 * ' ' + '1\n')

                            if 'no' in flag:
                                arqAtp.write('/OUTPUT\n')

                            flag = 'done'

                        if 'C  dT  >< Tmax >< Xopt >< Copt ><Epsiln>' in linha:
                            flag = 'top'

                        if '/BRANCH' in linha:
                            flag = 'branch'

                        if '/SWITCH' in linha:
                            flag = 'switch'

                        if '/OUTPUT' in linha and flag != 'done':
                            arqAtp.write('/SWITCH\n')
                            flag = 'no_switch'
                            continue

                        arqAtp.write(linha)

                tempBatch.write(batchConst[0] + tempAtp.name + batchConst[1])

        tempBatch.close()

        # Aqui roda os .atps e faz a leitura dos valores de curto-circuito

        run(str(batchPath), shell = True)

    for arquivo in atpWork.glob('*f_*.lis'):

        flag = 0

        with Path(arquivo).open('r') as lisFile:

            for linha in lisFile:

                if 'CORRESPONDE A BARRA' in linha:
                    barra = int(linha.split('|')[1].split()[4])

                if flag == 1:
                    vals = linha.split()
                    if 'CURTOA' in vals[1]:
                        faseA = [float(vals[4]) / sqrt(2) / 1e3, abs(tan(radians(float(vals[5]))))]
                    elif vals[2] == 'Open':
                        valsCurto[barra]['1F']['ATP'] = faseA
                        try:
                            valsCurto[barra]['1F']['DIFF'] = [(1 - valsCurto[barra]['1F']['ANAFAS'][0] / faseA[0]) * 100,
                                                          (1 - valsCurto[barra]['1F']['ANAFAS'][1] / faseA[1]) * 100]
                        except(ZeroDivisionError):
                            valsCurto[barra]['1F']['DIFF'] = [99999, 99999]
                        break
                    else:
                         faseB = [float(vals[4]) / sqrt(2) / 1e3, abs(tan(radians(float(vals[5]))))]
                         flag = 2

                if 'Output for steady-state phasor switch currents' in linha:
                    lisFile.readline()
                    flag = 1

                if flag == 2:

                    valsCurto[barra]['3F']['ATP'] = [(faseA[0] + faseB[0] + float(vals[4]) / sqrt(2) / 1e3) / 3,
                                                    (faseA[1] + faseB[1] + abs(tan(radians(float(vals[5]))))) / 3]

                    try:
                        valsCurto[barra]['3F']['DIFF'] = [(1 - valsCurto[barra]['3F']['ANAFAS'][0] / valsCurto[barra]['3F']['ATP'][0]) * 100,
                                                      (1 - valsCurto[barra]['3F']['ANAFAS'][1] / valsCurto[barra]['3F']['ATP'][1]) * 100]
                    except(ZeroDivisionError):
                        valsCurto[barra]['3F']['DIFF'] = [99999, 99999]

    relaRncc = arqPaths['cwd'].resolve() / Path(str(arqPaths['Ana'].resolve().stem) + '-rncc.rel')

    # Begin sorting of differences

    diffsCresc = []
    for barra in valsCurto:
        if valsCurto[barra]['3F'].get('ATP') != None:
            diffsCresc.append([abs(valsCurto[barra]['3F']['DIFF'][0]), barra])
    diffsCresc = sorted(diffsCresc)

    # Começa escrita no relatório

    with relaRncc.open('w') as rela:

        rela.write('{:^33}  {:^35}  {:^35}\n'.format(
                                'IDENTIFICACAO', 'TRIFASICO', 'MONOFASICO'))

        rela.write('{:^5}  {:^12}  {:^5}  {:^5}  '.format(
                                                'NUM.', 'NOME', 'ATP', 'VBAS'))

        rela.write('{:^5} {:^5} {:^5} {:^5} {:^5} {:^5}  '.format(
                                        'kA', 'X/R', 'kA', 'X/R', '%', '%'))

        rela.write('{:^5} {:^5} {:^5} {:^5} {:^5} {:^5}\n'.format(
                                        'kA', 'X/R', 'kA', 'X/R', '%', '%'))

        rela.write(35 * ' ')

        rela.write('{:^11} {:^11} {:^11}  {:^11} {:^11} {:^11}\n'.format(
                    'ANAFAS', 'ATP', 'DIFERENÇA', 'ANAFAS', 'ATP', 'DIFERENÇA'))

        for diff in diffsCresc:

            barra = diff[1]

            if valsCurto[barra]['3F'].get('ATP') != None:
                rela.write('{!s:>5.5}  {:<12.12}  {:<5.5}  {!s:<5.5}  '.format(
                                                        barra,
                                                        dbar.get_nomeAna(barra),
                                                        dbar.get_nomeAtp(barra),
                                                        dbar.get_vBase(barra)))
                rela.write('{!s:>5.5} {!s:>5.5} {!s:>5.5} {!s:>5.5} {!s:>5.5} {!s:>5.5}  '.format(
                                        valsCurto[barra]['3F']['ANAFAS'][0],
                                        valsCurto[barra]['3F']['ANAFAS'][1],
                                        valsCurto[barra]['3F']['ATP'][0],
                                        valsCurto[barra]['3F']['ATP'][1],
                                        valsCurto[barra]['3F']['DIFF'][0],
                                        valsCurto[barra]['3F']['DIFF'][1]))
                rela.write('{!s:>5.5} {!s:>5.5} {!s:>5.5} {!s:>5.5} {!s:>5.5} {!s:>5.5}\n'.format(
                                        valsCurto[barra]['1F']['ANAFAS'][0],
                                        valsCurto[barra]['1F']['ANAFAS'][1],
                                        valsCurto[barra]['1F']['ATP'][0],
                                        valsCurto[barra]['1F']['ATP'][1],
                                        valsCurto[barra]['1F']['DIFF'][0],
                                        valsCurto[barra]['1F']['DIFF'][1]))


    return valsCurto






def make_Rela(relaBuffer, arqPaths):
    """Função para escrever o relatório. Ela é chamada toda vez que o buffer
    relaBuffer é escrito. O que será escrito no relatório dependerá do que está
    no buffer.
    O buffer é um tuple com uma string para selecionar o texto e um número
    arbitrário de elementos, de forma a fornecer a essa função as informações
    necessárias para escrever o relatório.
    relaBuffer = 'str' + *args
    O conjunto dos possíveis buffers é:
    'Ana' + binário indicando o sucesso da leitura do arquivo - Referente à lei-
    tura do arquivo .ANA com os dados de curto-circuito saídos do Anafas/SAPRE.
    'rncc' - Referente à emissão de Relatório de Níveis de Curto-circuito pelo
    Anafas
    'barras' - Referente à opção do usuário de apenas emitir as barras de fron-
    teira do equivalente .ANA.
    'atp' + binário indicando sucesso da leitura do arquivo - Referente à
    leitura do arquivo de texto com os nomes de barras de fronteira a serem uti-
    lizadas no ATP.
    'diff' + diferença no número de barras entre o arquivo .ANA e o de texto com
    os nomes de barras de fronteira do ATP.
    'src' - Referente a requisição de um cartão /SOURCE para o ATP com as fontes
    das barras de fronteira.
    'equi' + conjunto de barras equivalentes + conjunto de circuitos equivalen-
    tes - Referente à impressão no relatório dos dados de barras do equivalente.

    Essa função chama o arquivo textos.py, que possui os textos padrão para escre
    ver no relatório.
    """

    rela = arqPaths['cwd'] / Path(arqPaths['Ana'].name[0:-4] + '-relatorio.rel')

    if 'welcome' in relaBuffer:

        #Compõe o nome do arquivo de relatório a partir do nome do arquino .ANA
        rela = rela.open('w')
        data = datetime.datetime.now(GMT1())
        rela.write(textos.texto['welcome'].format(VERSION, data.day, data.month, data.year, data.hour, data.minute))

    else: rela = rela.open('a')


    if 'Ana' in relaBuffer[0]:
        if not relaBuffer[1]:
            rela.write(textos.texto['relaErroArq'].format(arqPaths['Ana'].absolute()))
        else: rela.write(textos.texto['relaArq'].format(arqPaths['Ana'].resolve()))

    if 'Negs' in relaBuffer[0]:
        rela.write(textos.texto['Negs'])
        rela.write('===EQUIVALENTES===\n')
        for neg in relaBuffer[1]['equiv']:
            rela.write(str(neg[0]) + ' - ' + str(neg[1]) + '\n')
        rela.write('\n===REDE INTERNA===\n')
        for neg in relaBuffer[1]['interna']:
            rela.write(str(neg[0]) + ' - ' + str(neg[1]) + '\n')
        rela.write('\n')

    if 'Rep' in relaBuffer[0]:
        rela.write(textos.texto['Repetido'])
        for barra in relaBuffer[1]:
            rela.write(barra)
        rela.write('\n\n')

    if 'rncc' in relaBuffer[0]:
        rela.write(textos.texto['relaRncc'].format(arqPaths['cwd'].resolve() / Path(str(arqPaths['Ana'].resolve().stem) + '-rncc.rel')))

    if 'barras' in relaBuffer[0]:
        barrasEquiv = relaBuffer[2].get_equiNodes()[1:]
        rela.write(textos.texto['relaBarra'].format(len(barrasEquiv)))

        rela.write('{0:^6s} {1:<11s} {2:^6s}\n'\
            .format(*textos.texto['cabecalho'].split()))
        for barra in barrasEquiv:
            rela.write('{0:^6d} {1:<12s} {2!s:^6}\n'.format(barra, relaBuffer[1].get_nomeAna(barra), relaBuffer[1].get_vBase(barra)))

    if 'atp' in relaBuffer[0]:
        if not relaBuffer[1]:
            rela.write(textos.texto['relaErroArq'].format(arqPaths['Nomes'], Path.cwd()/arqPaths['Nomes']))
        else: rela.write(textos.texto['relaArq'].format(arqPaths['Nomes'].name))

    if 'miss' in relaBuffer[0]:
        if relaBuffer[2] == 'equi':
            rela.write(textos.texto['relaErroMiss'].format('Equivalente'))
        else:
            rela.write(textos.texto['relaErroMiss'].format('Interna'))
        for barra in relaBuffer[1]:
            rela.write(str(barra[0]) + ' - ' + barra[1] +  ' - ' + barra[2] + '\n')
        rela.write('\n')

    if 'src' in relaBuffer[0]:
        rela.write(textos.texto['relaSrc'].format(arqPaths['cwd'].resolve() / Path(str(arqPaths['Ana'].stem) + '-source.lib')))

    if 'inner' in relaBuffer[0]:
        rela.write(textos.texto['Inner'].format(relaBuffer[1]))

    if 'equi' in relaBuffer[0]:
        fontes = set()
        for barra in relaBuffer[2].get_equiNodes():
            fontes.add(relaBuffer[1].get_nomeGerAtp(barra))
        rela.write(textos.texto['relaEqui'].format(Path.cwd()/arqPaths['cwd'].resolve() / Path(str(arqPaths['Ana'].stem) + '-equivalentes.lib'), len(relaBuffer[2].get_equiNodes()), len(fontes)-1))
        rela.write('{:^6}{:^15}{:^10}{:^10}\n'.format(*textos.texto['cabecalhoF'].split()))
        for barra in relaBuffer[2].get_equiNodes():
            rela.write('{:^6d}{:^15}{:^11}{:^11}\n'.format(barra, relaBuffer[1].get_nomeAna(barra), relaBuffer[1].get_nomeAtp(barra), relaBuffer[1].get_nomeGerAtp(barra)))




class relaWatcher():
    """Classe Descritor para monitorar o andamento do relatório, com o uso da
    função Property.
    Quando se altera o valor do buffer bufferRela, ele já chama a função que
    escreve o relatório. Assim, a escrita do relatório é dinâmica."""

    def __init__(self, make_Rela, arqPaths):
        self.make_Rela = make_Rela
        self.arqPaths = arqPaths

    def setter(self, value):
        self._relaBuffer = value
        self.make_Rela(self._relaBuffer, self.arqPaths)
        self._relaBuffer = 0

    def setter2(self, value):
        if not value:
            exit()

    relaBuffer = property(fset=setter)
    runTime = property(fset=setter2)



class GMT1(datetime.tzinfo):
    "Faz o fuso-horário local"
    def utcoffset(self, dt):
        return datetime.timedelta(hours=-3)
    def dst(self, dt):
        return datetime.timedelta(0)


def paramsIniciais(file_json):

    # Ler arquivo JSON, remover comentários e salvar as opções do usuário
    opcoes = ''
    try:
        with open(file_json,'r') as f:
            for line in f:
                line = line.partition('#')[0]
                line.rstrip()
                opcoes += line
    except:
        print("Arquivo ",file_json," inválido!\nPrograma terminado!")
        exit()

    params = json.loads(opcoes)

    arqPaths = {'Ana' : '',
                'Nomes' : '',
                'cwd' : '',
                'Rncc' : '',
                'Atp' : '',
                'ATPcwd' : '',
                'initSource': ''}

    for param in params['caminhos']:
        if params['caminhos'][param]:
            arqPaths[param] = Path(params['caminhos'][param])

    arqPaths['cwd'] = Path.cwd()
    arqPaths['Ana'] = arqPaths['cwd'] / arqPaths['Ana']
    arqPaths['Nomes'] = arqPaths['cwd'] / arqPaths['Nomes']
    arqPaths['Rncc'] = arqPaths['cwd'] / arqPaths['Rncc']
    arqPaths['Atp'] = arqPaths['ATPcwd'] / arqPaths['Atp']
    arqPaths['initSource'] = arqPaths['cwd'] / arqPaths['initSource']

    base = params['base']

    return arqPaths, base

def criaCasos(arqPaths):
    "Cria os arquivos de determinada manobra. Já inclui .lib estatístico e demais coisas ??????"

    fixaOpc = {'energ':0, 'relig':0, 'rmono':0}
    batch = Path('rodaTodos.bat').open('w')
    lista_saidas = []

    opcoesPadrao = {
    'energ':{
        'delta':'1.E-6',
        'tmax':'0.3',
        'tChaves':['0.01'],
        'disp':'0.00125'
        },
    'relig':{
        'delta':'1.E-6',
        'tmax':'0.8',
        'tChaves':['0.01', '0.15', '0.065', '0.5'],
        'disp':'0.00125'
        },
    'rmono':{
        'delta':'1.E-6',
        'tmax':'0.8',
        'tChaves':['0.01', '0.15', '0.065', '0.5'],
        'disp':'0.00125'
        }
    }

    for arq in arqPaths['manobras'].open('r'):

        # Identifica os nomes dos nós das extremidades da linha
        extr = [arq.strip()[6:11], arq.strip()[12:17]]

        # Detecta o tipo de manobra
        arquivo = Path(arq.strip())
        sim = arquivo.name[:5].lower()

        try:
            caso = arquivo.open('r', encoding='latin1')
        except(FileNotFoundError):
            print('{}\nArquivo {} não encontrado!\n{}\n'.format('*'*50, arquivo.name, '*'*50))
            time.sleep(2)
            continue


        # Detalhamento de tChaves:
        # Para ENER: tempo de fechamento
        # Para RELIG: tempo de aplicação da falta,
        # delay para abertura do DJ de RECEP após falta,
        # delay para remoção da falta (com sucesso) após abertura do RECEP e
        # delay para fechamento estatístico de EMISS após abertura de RECEP

        fases = ['A', 'B', 'C', 'A']
        nodes = ['RECEP', 'EMISS', 'MEIOL']

        try:
            pers = Path('tempos').open('rb')
            opcoes = pk.load(pers)
            pers.close()
            pers = Path('tempos').open('wb')
        except:
            pers = Path('tempos').open('wb')
            opcoes = deepcopy(opcoesPadrao)


        if sim == 'energ':
            tipo = 'ENERGIZAÇÃO'
            saidaNomes = [arquivo.stem + '_.atp', arquivo.stem + '_1F.atp']
            
        elif sim == 'relig':
            tipo = 'RELIGAMENTO TRIPOLAR'
            saidaNomes = [arquivo.stem + '_CS.atp', arquivo.stem + '_SS.atp']

        elif sim == 'rmono':
            tipo = 'RELIGAMENTO MONOPOLAR'
            saidaNomes = [arquivo.stem + '_.atp']

        saidas = []

        for saida in saidaNomes:
            saidas.append(Path(saida).open('w'))
            lista_saidas.append(saida)


        # Confirma com o usuário
        if fixaOpc[sim]:
            troca = ''
        else:
            troca = 'dummy'
            print('\n{}\nArquivo {}\nCaso de {}\n{}\nOs parâmetros considerados são:\n'.format('-'*40, arquivo, tipo, '-'*40))

        while(troca):
            print('{}: {}\n{}: {}'.format(
                '1) Timestep', opcoes[sim]['delta'],
                '2) Tempo de Simulação', opcoes[sim]['tmax']))
            if sim == 'relig' or sim == 'rmono':
                print('{}: {}\n{}: {}\n{}: {}\n{}: {}\n{}: {}\n'.format(
        '3) Instante de aplicação da falta',
        opcoes[sim]['tChaves'][0],
        '4) Tempo para abertura do DJ próximo à falta',
        opcoes[sim]['tChaves'][1],
        '5) Delay para remoção da falta (com sucesso) após abertura do DJ',
        opcoes[sim]['tChaves'][2],
        '6) Delay para fechamento estatístico do DJ EMISS após abertura do RECEP',
        opcoes[sim]['tChaves'][3],
        '7) Tempo de dispersão',
        opcoes[sim]['disp']))
            
                troca = input('Digite o número do parâmetro que deseja mudar [1-7], ENTER para nenhum, * para restaurar padrões,\
                    = para fixar essa configuração para o resto das manobras deste tipo\n')

            elif sim == 'energ':
                print('{}: {}\n{}: {}\n'.format(
        '3) Instante de tempo da manobra',
        opcoes[sim]['tChaves'][0],
        '4) Tempo de dispersão',
        opcoes[sim]['disp']))
                troca = input('Digite o número do parâmetro que deseja mudar [1-4], ENTER para nenhum, * para restaurar padrões,\
                    = para fixar essa configuração para o resto das manobras deste tipo\n')

            if not troca:
                break

            elif troca == '1':
                opcoes[sim]['delta'] = input('Timestep: ')

            elif troca == '2':
                opcoes[sim]['tmax'] = input('Tempo de Simulação: ')

            elif troca == '3':
                if sim == 'ener':
                    opcoes[sim]['tChaves'][0] = input('Instante de tempo da manobra: ')
                else:
                    tChaves[sim]['tChaves'][0] = input('Instante de aplicação da falta: ')

            elif troca == '4':
                if sim == 'ener':
                    opcoes[sim]['disp'] = input('Tempo de dispersão: ')
                else:
                    tChaves[sim]['tChaves'][1] = input('Tempo para abertura do DJ: ')

            elif troca == '5':
                try:
                    opcoes[sim]['tChaves'][2] = input('Delay para remoção da falta: ')
                except(KeyError):
                    print('Valor inválido\n')

            elif troca == '6':
                try:
                    opcoes[sim]['tChaves'][3] = input('Delay para fechamento estatístico: ')
                except(KeyError):
                    print('Valor inválido\n')

            elif troca == '7':
                try:
                    opcoes[sim]['disp'] = input('Tempo de dispersão: ')
                except(KeyError):
                    print('Valor inválido\n')

            elif troca == '*':
                opcoes[sim] = deepcopy(opcoesPadrao[sim])

            elif troca == '=':
                fixaOpc[sim] = 1
                troca = ''

            else:
                print('Valor inválido\n')

        pk.dump(opcoes, pers)


        flag = 'misc'
        for linha in caso:

            # Detecta os cartões de MISCELANEOUS DATA
            try:
                if flag == 'misc':
                    if float(linha[:8]):
                        tempLinha = caso.readline()
                        saidas[0].write('C MISCELANEOUS DATA ORIGINAIS\nC ' + linha + 'C ' + tempLinha)
                        saidas[0].write('{:>8}{:>8}'.format(opcoes[sim]['delta'], opcoes[sim]['tmax']) + linha[16:])
                        saidas[0].write(tempLinha[:64] + '{:>8}'.format('5') + tempLinha[72:])
                        # Escreve cartão estatístico
                        saidas[0].write('{:>8}{:>8}{:>8}{:>8}{:>8}{:>8}{:>8}{:16}{:>8}\n'.format(
                            '1', '1', '0', '0', '1', '-1', '', '', '0'))
                        flag = 'cartao'
                        continue
            except(ValueError):
                pass

            # Detecta o primeiro cartão /"Alguma coisa" para inserir os cartões particulares do caso
            if flag == 'cartao':
                if linha[0] == '/':

                    # Insere as chaves para medição de tensões temporárias
                    saidas[0].write('C ===== CHAVES PARA MEDIÇÃO DE TENSÕES TEMPORÁRIAS ======\n/SWITCH\n')
                    for node in nodes:
                        for fase in fases[:-1]:
                            saidas[0].write('  {}{}TP{}{}{:>10}{:>10}{}0\n'.format(
                                node, fase, node[:3], fase, '.1', '1.e3', 45*' '))

                    # Insere as variaveis de saída
                    saidas[0].write('C ===== CONFIGURACAO DAS VARIAVEIS DE SAIDA =====\n')
                    saidas[0].write('/OUTPUT')
                    for node in nodes:
                        saidas[0].write('\n  ')
                        for fase in fases[:-1]:
                            saidas[0].write('TP{}{}{}{}'.format(node[:3], fase, node, fase))
                    for node in nodes:
                        saidas[0].write('\n-5')
                        for f in range(3):
                            saidas[0].write('TP{}{}TP{}{}{}{}{}{}'.format(
                                node[:3], fases[f], node[:3], fases[f + 1], node, fases[f], node, fases[f + 1]))


                    # Insere falta no nó RECEP
                    saidas[0].write('\nC ===== FALTA NO FIM DA LINHA =====\nC /BRANCH\n')
                    if sim == 'energ':
                        saidas[0].write('C   CURTOA{}{:>6}{}1\n'.format(' '*18, '1.e-6', 47*' '))
                        saidas[0].write('C /SWITCH\n')
                        saidas[0].write('C   RECEPACURTOA{:>10}{:>10}{}0\n'.format('-1.', '1.e3', 45*' '))
                    else:
                        saidas[0].write('  CURTOA{}{:>6}{}1\n'.format(' '*18, '1.e-6', 47*' '))
                        saidas[0].write('/SWITCH\n')
                        saidas[0].write('  RECEPACURTOA{:>10}{:>10.9g}{}0\n'.format(
                            opcoes[sim]['tChaves'][0],
                            float(opcoes[sim]['tChaves'][0]) + float(opcoes[sim]['tChaves'][1]) + 0.02 + float(opcoes[sim]['tChaves'][2]),
                            45*' '))

                    # Insere o cartão estatístico
                    saidas[0].write('C ===== CARTÃO PARA TABELAMENTO ESTATÍSTICO =====\n')
                    saidas[0].write('/STATISTICS\n')
                    saidas[0].write('STATISTICS DATA                2    0.05')

                    baseVf = 230e3 * sqrt(2/3)
                    baseVl = 230e3 * sqrt(2)
                    baseI = 1000
                    baseJ = 1000
                    findBuff = []
                    for node in nodes:
                        saidas[0].write('\n 0{:>12.5f}'.format(baseVf))
                        findBuff.append('\n 0{:12}'.format(''))
                        for fase in fases[:-1]:
                            toWrite = '{}{}{}'.format(node, fase, ' '*6)
                            saidas[0].write(toWrite)
                            findBuff.append(toWrite)
                        findBuff.append('\nDISK')

                        saidas[0].write('\n-1{:>12.5f}'.format(baseVl))
                        findBuff.append('\n-1{:12}'.format(''))
                        for f in range(3):
                            toWrite = '{}{}{}{}'.format(node, fases[f], node, fases[f + 1])
                            saidas[0].write(toWrite)
                            findBuff.append(toWrite)
                        findBuff.append('\nDISK')

                    # saidas[0].write('\n-2{:>12}'.format(base))
                    # findBuff.append('\n-2{:>12}'.format(''))
                    # for fase in fases[:-1]:
                    #     toWrite = '{}{}EMISS{}'.format(extr[0], fase, fase)
                    #     saidas[0].write(toWrite)
                    #     findBuff.append(toWrite)
                    # findBuff.append('\nDISK')

                    for node in nodes[:-1]:
                        saidas[0].write('\n-4{:>12.5f}'.format(baseJ))
                        findBuff.append('\n-4{:12}'.format(''))
                        for fase in fases[:-1]:
                            toWrite = '{}{}{}'.format(node, fase, ' '*6)
                            saidas[0].write(toWrite)
                            findBuff.append(toWrite)
                        findBuff.append('\nDISK')

                    saidas[0].write('\nFIND')
                    for find in findBuff:
                        saidas[0].write(find)
                    saidas[0].write('\nQUIT\n')

                    saidas[0].write('C ===== AGORA INICIA OS CARTÕES DO CASO ORIGINAL =====\n')

                    flag = 'PR'                    

            if flag == 'blank':
                if linha.startswith('BEGIN'):
                    saidas[0].write('BLANK STATISTICS\n')
                    flag = 'done'

            if flag == 'chaves':
                # Procura pela chave determinística do emissor
                match = search('^..{}.EMISS.'.format(extr[0]), linha)
                if match:
                    if sim == 'energ':
                        saidas[0].write('76{}{:>10}{:>10}{}STATISTICS{}1\n'.format(
                            match.group()[2:], opcoes[sim]['tChaves'][0], opcoes[sim]['disp'], ' '*20, ' '*15))
                        continue

                    elif sim == 'relig':
                        saidas[0].write('{}{:>10}{:>10.9g}{}1\n'.format(
                            match.group(),
                            '-1.',
                            float(opcoes[sim]['tChaves'][1]) + 0.02 + float(opcoes[sim]['tChaves'][0]),
                            ' '*45))
                        saidas[0].write('76{}{:>10.9g}{:>10}{}STATISTICS{}1\n'.format(
                            match.group()[2:],
                            0.02 + float(opcoes[sim]['tChaves'][0]) + float(opcoes[sim]['tChaves'][1]) + float(opcoes[sim]['tChaves'][3]),
                            opcoes[sim]['disp'],
                            ' '*20,
                            ' '*15))
                        continue

                    elif sim == 'rmono':
                        if match.group()[-1] == 'A':
                            saidas[0].write('{}{:>10}{:>10.9g}{}1\n'.format(match.group(),
                                '-1.',
                                float(opcoes[sim]['tChaves'][1]) + 0.02 + float(opcoes[sim]['tChaves'][0]),
                                ' '*45))
                            saidas[0].write('76{}{:>10.9g}{:>10}{}STATISTICS{}1\n'.format(
                                match.group()[2:],
                                0.02 + float(opcoes[sim]['tChaves'][0]) + float(opcoes[sim]['tChaves'][1]) + float(opcoes[sim]['tChaves'][3]),
                                opcoes[sim]['disp'],
                                ' '*20, ' '*15))
                        else:
                            saidas[0].write('{}{:>10}{:>10}{}1\n'.format(match.group(), '-1.', '100.', ' '*45))
                        continue

                # Procura pela chave determinística do receptor
                match = search('^..RECEP.{}.'.format(extr[1]), linha)
                if match:
                    if sim == 'energ':
                        saidas[0].write('{}{:>10}{:>10}{}0\n'.format(match.group(), '100.', '1000.', ' '*45))
                        continue
                    elif sim == 'relig':
                        saidas[0].write('{}{:>10}{:>10.9g}{}0\n'.format(match.group(), '-1.', float(opcoes[sim]['tChaves'][0]) + float(opcoes[sim]['tChaves'][1]), ' '*45))
                        continue
                    elif sim == 'rmono':
                        if match.group()[7] == 'A':
                            saidas[0].write('{}{:>10}{:>10.9g}{}0\n'.format(match.group(), '-1.', float(opcoes[sim]['tChaves'][0]) + float(opcoes[sim]['tChaves'][1]), ' '*45))
                        else:
                            saidas[0].write('{}{:>10}{:>10}{}1\n'.format(match.group(), '-1.', '100.', ' '*45))
                        continue

                if linha.startswith('/'):
                    flag = 'blank'
                    
            if flag == 'PR':
            # Insere monitorações de energia nos PRs
            # Considera-se que os PRs estão conectados nos nós RECEP e EMISS
                if linha[:2] == '99':
                    saidas[0].write(linha[:-2] + '4\n')
                    continue

                if linha[:7] == '/SWITCH':
                    flag = 'chaves'


            saidas[0].write(linha)

        caso.close()
        saidas[0].close()


        # Cria o caso com curto no fim da linha (Energização) e caso sem sucesso (religamento)
        if sim == 'energ' or sim == 'relig':
            with Path(saidaNomes[0]).open('r') as saida:
                for linha in saida:

                    if sim == 'energ':
                        if search('^C ===== FALTA', linha):
                            saidas[1].write(linha)
                            for n in range(4):
                                saidas[1].write(saida.readline()[2:])
                            continue

                    elif sim == 'relig':
                        if search('..RECEPACURTOA', linha):
                            saidas[1].write('{}{:>10}{}'.format(linha[:24], '100.', linha[34:]))
                            continue

                    saidas[1].write(linha)
                saidas[1].close()


    # Escreve o arquivo batch
    batch.write('echo off\nsetlocal EnableDelayedExpansion\n')
    batch.write('SET GNUDIR={}\\\nSET DIREXE={}\n'.format(
        arqPaths['GNUDIR'],
        arqPaths['tpbigDir']))
    atps = [Path(arq).stem for arq in lista_saidas]
    for n in range(len(atps)):
        batch.write('set arq[{}]={}\n'.format(n, atps[n]))
    batch.write('for /L %%i in (0,1,{:d}) do (\n'.format(len(atps) - 1))
    batch.write('    %DIREXE%\\{} disk !arq[%%i]!.atp s -r > !arq[%%i]!.log 2>&1\n    MD !arq[%%i]!\n'.format(arqPaths['tpbig']))
    batch.write('    MOVE SHOT*.DAT .\!arq[%%i]!\n    MOVE !arq[%%i]!.* .\!arq[%%i]!\n    )\n')
    batch.write('\nDEL *.BIN\nDEL *.BAK\nDEL *.46\nDEL *.6\nREM shutdown -s now')

    batch.close()




def main():

    # Trata dos argumentos de linha de comando
    argumnt = args_Handler()

    # Leitura do arquivo JSON e das opções do usuário
    # Abrir, ler, eliminar comentários e salvar as opções de entrada
    arqPaths, base = paramsIniciais(argumnt.args.json)

    # Experimentação da nova função de configurar automaticamente as manobras
    if argumnt.args.m:
        criaCasos(arqPaths)
        exit()

    #Instancia a classe de monitoramento do status do relatório. Conforme os
    # processos vão avançando, o relatório vai sendo escrito.
    relaWatch = relaWatcher(make_Rela, arqPaths)

    # Imprime a mensagem de boas-vindas
    data = datetime.datetime.now(GMT1())
    print(textos.texto['welcome'].format(VERSION, data.day, data.month, data.year, data.hour, data.minute))
    relaWatch.relaBuffer = ('welcome',)

    # Verifica a existência do arquivo .ANA
    try:
        arqPaths['Ana'].resolve()
    except(FileNotFoundError):
        relaWatch.relaBuffer = ('Ana',0)
        relaWatch.runTime = 0

    relaWatch.relaBuffer = ('Ana',1)

    # Inicia a execução das operações e obtenção de dados

    # Obtém a lista de barras de fronteira e os circuitos equivalentes
    #conectados a elas do arquivo .ANA
    dbar = get_DBAR(codecs.open(str(arqPaths['Ana'].resolve()), mode="r", encoding="iso-8859-1"), Nodes())

    dlin = get_EQUIV(codecs.open(str(arqPaths['Ana'].resolve()), mode="r", encoding="iso-8859-1"), Branches(), dbar)

    # A seguir é feita a seleção do modo de operação do programa, de acordo com
    #os argumentos que o usuário entrou na linha de comando.

    # b  Impressão de lista de barras de fronteira do arquivo .ANA
    if argumnt.args.b:
        relaWatch.relaBuffer = ('barras', dbar, dlin)

    else:
        try:
            # missing guarda os nomes de nohs do ATP faltantes.
            # repet guarda o nome de nohs repetidos.
            autoEquiv, autoIntern = getAtpNames(arqPaths, dbar, dlin, base)
            repet = dbar.get_repATP()

        except(FileNotFoundError):
            relaWatch.relaBuffer = ('atp', 0)
            relaWatch.runTime = 0

        if repet:
            relaWatch.relaBuffer = ('Rep', repet)

        if autoEquiv:
            relaWatch.relaBuffer = ('miss', autoEquiv, 'equi')

        # R  Comparação de níveis de curto-circuito do ANAFAS com os do ATP
        if argumnt.args.R:
            valsCurto = compCurto(arqPaths, dbar, jump=0)
            relaWatch.relaBuffer = ('rncc', valsCurto)

        # Rj Idem ao Comando R, porém não roda novamente os curto-circuitos do ATP
        if argumnt.args.Rj:
            valsCurto = compCurto(arqPaths, dbar, jump=1)
            relaWatch.relaBuffer = ('rncc', valsCurto)

    if argumnt.args.s:
        makeSource(arqPaths, dbar)
        relaWatch.relaBuffer = ('src',)

    if argumnt.args.i:
        makeLib(arqPaths, dlin, dbar, inner = 1)
        relaWatch.relaBuffer = ('inner', arqPaths['cwd'].resolve() / Path(str(arqPaths['Ana'].stem) + '-interna.lib'))
        if autoIntern:
            relaWatch.relaBuffer = ('miss', autoIntern, 'inner')

    if argumnt.args.e:
        makeLib(arqPaths, dlin, dbar, inner = 0)
        relaWatch.relaBuffer = ('equi', dbar, dlin)

    # Verifica se há alguma valor negativo de parametro, e alerta o usuario
    if dlin.negs:
        relaWatch.relaBuffer = ('Negs', dlin.negs)

    # FIM DA EXECUÇÃO
    relaWatch.relaBuffer = ('fim',)

main()
